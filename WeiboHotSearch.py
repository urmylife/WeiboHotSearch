import requests
from bs4 import BeautifulSoup
import time
import openpyxl as xl


def get_html_text(url):
    headers = {
        'Cookie': 'WBStorage=6b696629409558bc|undefined;Ugrow-G0=d52660735d1ea4ed313e0beb68c05fc5;wb_view_log_7125808863=2560*14401;login_sid_t=3ad62385718b3b5b1f92ac972379c12a;cross_origin_proto=SSL;ULV=1560774064175:1:1:1:2508389041312.6465.1560774064170:;YF-V5-G0=125128c5d7f9f51f96971f11468b5a3f;YF-Page-G0=761bd8cde5c9cef594414e10263abf81|1560774070|1560774070;_s_tentry=passport.weibo.com;SINAGLOBAL=2508389041312.6465.1560774064170;Apache=2508389041312.6465.1560774064170;wb_view_log=2560*14401;SUHB=0vJYvuxB_LFc8r;SUBP=0033WrSXqPxfM725Ws9jqgMF55529P9D9W5UBHJaQozGsMlSa0_yQKpd5JpX5K2hUgL.FoMpeo-RehnRSoe2dJLoI7vWIPiPgH2LxK-LBo5LBo2t;ALF=1592310065;SSOLoginState=1560774066;SCF=AqXCTaf0OiHe7c5TAy3Ozm2IvbB4YZEjzGfCIotUE62mNpFvjgMpN9PYSJIdju5-GORERsB_-UGmARMFDAxLMhA.;SUB=_2A25wA_XlDeThGeFP6VcZ8CbEzT-IHXVTeWAtrDV8PUNbmtAKLWvfkW9NQTkdOjOcgSBwIu43ODBkrMLcDNwT_KIV;un=edx7pv7wb5@gggggs.fun;wvr=6;webim_unReadCount=%7B%22time%22%3A1560774074466%2C%22dm_pub_total%22%3A0%2C%22chat_group_pc%22%3A0%2C%22allcountNum%22%3A52%2C%22msgbox%22%3A0%7D',

        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/52.0.2743.116 Safari/537.36 Edge/15.15063'}
    headers = {
        'Cookie': 'SINAGLOBAL=9993033941305.139.1522647209174; UOR=ent.ifeng.com,widget.weibo.com,www.baidu.com; login_sid_t=8c241ba284c4fafac503b7731708194c; cross_origin_proto=SSL; _s_tentry=www.baidu.com; Apache=3596338825797.396.1584498701374; ULV=1584498701381:3:1:1:3596338825797.396.1584498701374:1577172939435; SUBP=0033WrSXqPxfM725Ws9jqgMF55529P9D9Whrfs8sHS6F1G9RYeCxry9d5JpX5K2hUgL.FoqEe0nc1hqfe0n2dJLoIEBLxKqL1KnLB-qLxKnL1K5LBKMLxKnLBKnL1h5LxK-L1K2L1h5t; ALF=1616034947; SSOLoginState=1584498948; SCF=AumouGoPo94zXv0L61TlPMbBAMj3lW0s2zLSTP47uuoxX9a8YEnxswoONVm1DNMuAruNaoTu6zEmGrYBVcCoKXY.; SUB=_2A25zdflUDeRhGeBM6FoX-CjJyDSIHXVQA22crDV8PUNbmtAKLW2skW9NROXyt55j1kU09mbjwmVf1v1Yah1NBd1Z; SUHB=0WGMXjb3rIKb_c; wvr=6; webim_unReadCount=%7B%22time%22%3A1584500158087%2C%22dm_pub_total%22%3A4%2C%22chat_group_client%22%3A0%2C%22allcountNum%22%3A46%2C%22msgbox%22%3A0%7D; WBStorage=42212210b087ca50|undefined',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.149 Safari/537.36'}
    try:
        r = requests.get(url, headers=headers)
        r.raise_for_status()
        r.encoding = 'utf-8'
        print(r.text)
        return r.text
    except:
        return ""


def get_html_list(ilt, html, num):
    tplt = "{0:^10}\t{1:^4}\t{2:{4}^20}\t{3:^10}"
    print("{0:^16}\t{1:^12}\t{2:^15}\t{3:^25}".format("Time", "Rank", "Num", "Title", chr(12288)))
    soup = BeautifulSoup(html, 'html.parser')
    i = 1
    try:
        items = soup.find_all('td', class_='td-02')
        time_stamp = time.strftime('%Y/%m/%d %H:%M', time.localtime(time.time()))
        for item in enumerate(items[1:num+1]):
            num = item[1].find('span').text
            title = item[1].find('a').text
            rank = "第{}名".format(i)
            i += 1
            print(tplt.format(time_stamp, rank, num, title, chr(12288)))
            ilt.append([time_stamp, rank, num, title])
    except:
        ""


def save_html_list(input_path, output_path, ilt, num):
    wb = xl.load_workbook(input_path)
    sheet1 = wb['Sheet1']
    sheet1.cell(1, 1).value = "Time"
    sheet1.cell(1, 2).value = "Rank"
    sheet1.cell(1, 3).value = "Num"
    sheet1.cell(1, 4).value = "Title"
    for i in range(num):
        u = ilt[i]
        j = i + 2
        sheet1.cell(j, 1).value = u[0]
        sheet1.cell(j, 2).value = u[1]
        sheet1.cell(j, 3).value = u[2]
        sheet1.cell(j, 4).value = u[3]
    wb.save(output_path)


def main():
    uinfo = []
    input_path = r'D:\File\python\Weibo_Dynamic_Ranking\Weibo_Dynamic_Ranking.xlsx'
    output_path = r'D:\File\python\Weibo_Dynamic_Ranking\Weibo_Dynamic_Ranking.xlsx'
    num = 10
    html = get_html_text(url="https://s.weibo.com/top/summary?Refer=top_hot&topnav=1&wvr=6&display=0&retcode=6102")
    get_html_list(uinfo, html, num)
    save_html_list(input_path, output_path, uinfo, num)


main()
